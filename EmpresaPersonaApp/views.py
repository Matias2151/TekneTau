# EmpresaPersonaApp/views.py
from datetime import datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal
from django.db import transaction
from django.db.models import Count, F, OuterRef, Q, Sum, Prefetch
from django.db.models.expressions import Subquery
from django.db.models.functions import Coalesce, TruncMonth
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.contrib import messages  

from UsuariosApp.forms import LoginForm
from UsuariosApp.models import PasswordResetCode,UsuarioSistema

from .models import EmpresaPersona
from .forms import EmpresaPersonaForm
from DireccionApp.models import Direccion
from DireccionApp.forms import DireccionForm
from FacturacionApp.models import DetalleDoc, Documento
from ProyectoApp.models import Proyecto

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# =====================================
#   VISTA PRINCIPAL EMPRESA/PERSONA
# =====================================
@transaction.atomic
def empresa_clientes(request):
    """
    Vista principal de gestión de Empresa/Persona:
    - Lista todas las personas (clientes/proveedores)
    - Permite crear nuevas desde el modal.
    - Separa habilitados vs deshabilitados para usarlos en pestañas.
    """

    # Consultamos todas las personas con su dirección relacionada
    personas = EmpresaPersona.objects.select_related(
        "emppe_dire__regi", "emppe_dire__ciuda", "emppe_dire__comun"
    ).all()

    # Subgrupos por estado
    personas_habilitadas = personas.filter(emppe_est=True)
    personas_deshabilitadas = personas.filter(emppe_est=False)

    if request.method == "POST":
        # Formularios de creación
        form_p = EmpresaPersonaForm(request.POST)
        form_d = DireccionForm(request.POST)

        if form_p.is_valid() and form_d.is_valid():
            # Guardamos primero la dirección
            direccion = form_d.save()
            # Luego la persona, enlazando la dirección
            persona = form_p.save(commit=False)
            persona.emppe_dire = direccion
            persona.save()

            messages.success(
                request,
                f'Se ha creado el cliente/proveedor "{persona.emppe_nom}" correctamente.'
            )
            return redirect("empresa_clientes")

        # ❗Si hay errores, se vuelve a renderizar con los formularios con errores
        context = {
            # En la tabla mostraremos TODAS las personas y filtraremos por JS (pestañas)
            "personas": personas,
            "personas_habilitadas": personas_habilitadas,
            "personas_deshabilitadas": personas_deshabilitadas,
            "form_p": form_p,
            "form_d": form_d,
        }
        return render(request, "empresa_persona/empresa_clientes.html", context)

    # GET normal
    context = {
        "personas": personas,
        "personas_habilitadas": personas_habilitadas,
        "personas_deshabilitadas": personas_deshabilitadas,
        "form_p": EmpresaPersonaForm(),
        "form_d": DireccionForm(),
    }
    return render(request, "empresa_persona/empresa_clientes.html", context)


# ===============================
#      EDITAR PERSONA
# ===============================
@transaction.atomic
def editar_persona(request, pk):
    persona = get_object_or_404(EmpresaPersona, pk=pk)
    direccion = persona.emppe_dire

    if request.method == "POST":
        form_p = EmpresaPersonaForm(request.POST, instance=persona)
        form_d = DireccionForm(request.POST, instance=direccion)

        if form_p.is_valid() and form_d.is_valid():
            form_d.save()
            form_p.save()
            messages.success(
                request,
                f'Se ha actualizado la información de "{persona.emppe_nom}".'
            )
            return redirect("ver_persona", pk=persona.pk)

    # 🔥 NUEVO: renderizar solo el formulario si es GET (para modal)
    form_p = EmpresaPersonaForm(instance=persona)
    form_d = DireccionForm(instance=direccion)

    return render(
        request,
        "empresa_persona/editar_persona_form.html",
        {
            "form_p": form_p,
            "form_d": form_d,
            "persona": persona,
        }
    )

# ================================
#      INHABILITAR PERSONA
# ================================
@transaction.atomic
def eliminar_persona(request, pk):
    """
    'Elimina' lógicamente una EmpresaPersona.
    En realidad NO se borra de la BD, solo se marca emppe_est = False.
    """
    persona = get_object_or_404(EmpresaPersona, pk=pk)
    if request.method == "POST":
        persona.emppe_est = False
        persona.save(update_fields=["emppe_est"])

        messages.warning(
            request,
            f'Cliente/proveedor "{persona.emppe_nom}" fue inhabilitado correctamente.'
        )
        return redirect("empresa_clientes")

    return JsonResponse({"error": "Método no permitido"}, status=405)


# =================================
#      HABILITAR PERSONA
# ================================
@transaction.atomic
def habilitar_persona(request, pk):
    """
    Habilita nuevamente una EmpresaPersona previamente inhabilitada
    (emppe_est pasa a True).
    """
    persona = get_object_or_404(EmpresaPersona, pk=pk)
    if request.method == "POST":
        persona.emppe_est = True
        persona.save(update_fields=["emppe_est"])

        messages.success(
            request,
            f'Cliente/proveedor "{persona.emppe_nom}" fue habilitado correctamente.'
        )
        return redirect("empresa_clientes")

    return JsonResponse({"error": "Método no permitido"}, status=405)


# =========================
#       OTRAS VISTAS
# =========================

def index(request):
    """Redirige al inicio de sesión si no hay sesión activa."""
    if not request.session.get("user_role"):
        return redirect("login")
    return redirect("dashboard")

# =========================
#      LOGIN VIEW
# =========================

def login_view(request):
    if request.session.get("user_role"):
        return redirect("dashboard")

    form = LoginForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        username = form.cleaned_data["usuario"].strip()
        password = form.cleaned_data["password"]

        usuario = UsuarioSistema.objects.filter(username__iexact=username).first()

        if usuario is None:
            messages.error(request, "Usuario o contraseña incorrectos.")
            return render(request, "login/login.html", {"form": form})

        if not usuario.is_active:
            messages.error(
                request,
                "Tu cuenta está desactivada. Contacta al administrador."
            )
            return render(request, "login/login.html", {"form": form})

        if not usuario.check_password(password):
            codigo_temporal = (
                PasswordResetCode.objects.filter(
                    user=usuario, code=password, used=False
                )
                .order_by("-created_at")
                .first()
            )

            if not (codigo_temporal and codigo_temporal.is_valid()):
                messages.error(request, "Usuario o contraseña incorrectos.")
                return render(request, "login/login.html", {"form": form})

            codigo_temporal.used = True
            codigo_temporal.save(update_fields=["used"])

        request.session["user_id"] = usuario.id
        request.session["user_role"] = usuario.role
        request.session["username"] = usuario.username

        messages.success(
            request,
            "Ingreso exitoso. Te recomendamos actualizar tu contraseña si usaste una temporal.",
        )
        return redirect("dashboard")

    return render(request, "login/login.html", {"form": form})

def _clp0(x):
    # redondeo a entero CLP
    return int(Decimal(x).quantize(Decimal("1"), rounding=ROUND_HALF_UP))

# =========================
#     DASHBOARD VIEW
# =========================
def dashboard(request):
    hoy = timezone.localdate()

    # =========================
    # FILTRO MULTI-MES (GET)
    # =========================
    meses_sel_raw = request.GET.getlist("mes")  # ?mes=4&mes=5...
    meses_sel = [int(m) for m in meses_sel_raw if str(m).isdigit()]

    anio = request.GET.get("anio")
    anio = int(anio) if (anio and anio.isdigit()) else hoy.year

    # =========================
    # DOCUMENTOS BASE
    # =========================
    documentos = (
        Documento.objects
        .select_related("tipo_doc", "empresa", "proyecto")
        .prefetch_related("detalles__producto", "transacciones__tipo")
        .filter(docum_fecha_emi__year=anio)
        .order_by("-docum_fecha_emi", "-docum_id")
    )

    if meses_sel:
        documentos = documentos.filter(docum_fecha_emi__month__in=meses_sel)

    # =========================
    # MÉTRICAS
    # =========================
    docs_pendientes = 0
    docs_por_vencer = 0
    docs_atrasados  = 0

    total_ingresos = Decimal(0)
    total_egresos  = Decimal(0)

    saldo_pend_ing = Decimal(0)
    saldo_pend_egr = Decimal(0)

    por_vencer_list = []
    pendientes_list = []

    ESTADOS_PENDIENTE = {"PENDIENTE", "MITAD"}
    ventana_por_vencer = 7  # ajusta a 15/30 si quieres

    for doc in documentos:
        # -------- total/pagado/pendiente desde detalles --------
        doc_total = Decimal(0)
        doc_pagado = Decimal(0)

        for det in doc.detalles.all():
            precio = Decimal(det.producto.produ_bruto or 0) if det.producto else Decimal(0)
            cant   = Decimal(det.dedoc_cant or 0)

            doc_total += cant * precio

            pag_cant = Decimal(det.dedoc_pagado or 0)
            doc_pagado += pag_cant * precio

        doc_pendiente = doc_total - doc_pagado
        if doc_pendiente < 0:
            doc_pendiente = Decimal(0)

        # -------- tipo transacción del doc --------
        trans_tipo = None
        for t in doc.transacciones.all():
            tt = (t.tipo.tipo_trans or "").upper()
            if tt == "INGRESO":
                trans_tipo = "INGRESO"
                break
            elif tt == "EGRESO":
                trans_tipo = "EGRESO"

        # -------- totales globales --------
        if trans_tipo == "INGRESO":
            total_ingresos += doc_total
            saldo_pend_ing += doc_pendiente
        elif trans_tipo == "EGRESO":
            total_egresos += doc_total
            saldo_pend_egr += doc_pendiente

        # ======================================================
        # 🔥 PENDIENTES / POR VENCER / ATRASADOS (solo si hay saldo)
        # ======================================================
        estado = (doc.docum_estado or "").upper()
        if estado in ESTADOS_PENDIENTE and doc_pendiente > 0:
            docs_pendientes += 1

            # lista Pendientes/Mitad (segunda card)
            # dentro del if 0 <= dias_rest <= ventana_por_vencer:
            pendientes_list.append({
                "docum_num": doc.docum_num,
                "tipo_doc": getattr(doc.tipo_doc, "tidoc_tipo", ""),
                "cliente_id": doc.empresa.emppe_id if doc.empresa else None,
                "cliente": getattr(doc.empresa, "emppe_nom", "") if doc.empresa else "",
                "fecha_emi": doc.docum_fecha_emi,
                "pendiente": _clp0(doc_pendiente),
                "estado": doc.docum_estado,
                "css": "doc-ingreso" if trans_tipo == "INGRESO" else "doc-egreso",
            })



            # vencimiento: atrasados / por vencer
            if doc.docum_fecha_ven:
                if doc.docum_fecha_ven < hoy:
                    docs_atrasados += 1
                else:
                    dias_rest = (doc.docum_fecha_ven - hoy).days

                    # ✅ por vencer SOLO dentro de ventana
                    if 0 <= dias_rest <= ventana_por_vencer:
                        docs_por_vencer += 1

                        # dentro del if 0 <= dias_rest <= ventana_por_vencer:
                        por_vencer_list.append({
                            "docum_num": doc.docum_num,
                            "tipo_doc": getattr(doc.tipo_doc, "tidoc_tipo", ""),
                            "cliente_id": doc.empresa.emppe_id if doc.empresa else None,
                            "cliente": getattr(doc.empresa, "emppe_nom", "") if doc.empresa else "",
                            "fecha_ven": doc.docum_fecha_ven,
                            "dias_rest": dias_rest,
                            "total": _clp0(doc_total),
                            "pendiente": _clp0(doc_pendiente),
                            "estado": doc.docum_estado,
                            "css": "doc-ingreso" if trans_tipo == "INGRESO" else "doc-egreso",
                        })


    # orden listas
    por_vencer_list.sort(key=lambda x: x["dias_rest"])
    pendientes_list.sort(key=lambda x: (x["fecha_emi"] or hoy), reverse=True)

    utilidad_total = total_ingresos - total_egresos
    gasto_total = total_egresos

    # IVA gastado (si egresos es bruto)
    neto_egresos = (total_egresos / Decimal("1.19")) if total_egresos else Decimal(0)
    iva_gastado  = total_egresos - neto_egresos

    context = {
        "anio": anio,
        "meses_sel": meses_sel,

        "docs_pendientes": docs_pendientes,
        "docs_por_vencer": docs_por_vencer,
        "docs_atrasados": docs_atrasados,

        "saldo_pend_ing": _clp0(saldo_pend_ing),
        "saldo_pend_egr": _clp0(saldo_pend_egr),

        "utilidad_total": _clp0(utilidad_total),
        "gasto_total": _clp0(gasto_total),
        "iva_gastado": _clp0(iva_gastado),

        "por_vencer_list": por_vencer_list,
        "pendientes_list": pendientes_list,
    }

    return render(request, "login/dashboard.html", context)

# =========================
#   DATOS EN FORMATO JSON
# =========================
def obtener_personas_json(request):
    """
    Devuelve todas las empresas/personas en formato JSON.
    Ideal para usar con fetch() o DataTables en empresa_clientes.html.
    """
    personas = EmpresaPersona.objects.select_related(
        "emppe_dire__regi", "emppe_dire__ciuda", "emppe_dire__comun"
    ).all()

    data = []
    for p in personas:
        data.append({
            "id": p.emppe_id,
            "rut": p.emppe_rut,
            "nombre": p.emppe_nom,
            "alias": p.emppe_alias or "",
            "fono1": p.emppe_fono1,
            "fono2": p.emppe_fono2 or "",
            "mail1": p.emppe_mail1,
            "mail2": p.emppe_mail2 or "",
            "estado": "Activo" if p.emppe_est else "Inactivo",
            "situacion": p.emppe_sit,
            "direccion": {
                "calle": p.emppe_dire.dire_calle if p.emppe_dire else "",
                "num": p.emppe_dire.dire_num if p.emppe_dire else "",
                "otros": p.emppe_dire.dire_otros if p.emppe_dire else "",
                "region": p.emppe_dire.regi.regi_nom if p.emppe_dire and p.emppe_dire.regi else "",
                "ciudad": p.emppe_dire.ciuda.ciuda_nom if p.emppe_dire and p.emppe_dire.ciuda else "",
                "comuna": p.emppe_dire.comun.comun_nom if p.emppe_dire and p.emppe_dire.comun else "",
                "codigo_postal": p.emppe_dire.dire_cod_postal if p.emppe_dire else "",
            },
        })

    return JsonResponse(data, safe=False)

# ==================
#  VER PERSONA
# ==================
def ver_persona(request, pk):
    persona = get_object_or_404(EmpresaPersona, pk=pk)
    direccion = persona.emppe_dire

    if request.method == "POST":
        form_p = EmpresaPersonaForm(request.POST, instance=persona)
        form_d = DireccionForm(request.POST, instance=direccion)

        if form_p.is_valid() and form_d.is_valid():
            direccion = form_d.save()
            persona.emppe_dire = direccion
            persona.save()
            form_p.save()

            messages.success(request, "Datos del cliente y dirección actualizados correctamente.")
            return redirect("ver_persona", pk=persona.pk)
    else:
        form_p = EmpresaPersonaForm(instance=persona)
        form_d = DireccionForm(instance=direccion)

    proyectos = persona.proyectos.all()

    # Proyectos pendientes (ajusta estados según tus valores reales)
    proyectos_pendientes = proyectos.filter(
        Q(proye_estado__iexact="PENDIENTE") |
        Q(proye_estado__iexact="EN PROGRESO") |
        Q(proye_estado__iexact="EN_PROGRESO")
    ).count()

    # ==============================
    # DOCUMENTOS (QUERYSET BASE)
    # ==============================
    documentos = (
        persona.documentos
        .select_related("tipo_doc", "proyecto", "empresa", "forma_pago", "forma_pago__tipo_pago")
        .prefetch_related("detalles__producto", "transacciones__tipo")
        .order_by("-docum_fecha_emi", "-docum_id")
    )

    # ==============================
    # FILTRO POR FECHA (GET): DESDE / HASTA
    # ==============================
    desde_str = request.GET.get("desde", "").strip()
    hasta_str = request.GET.get("hasta", "").strip()

    desde = None
    hasta = None

    # parse seguro (formato input type="date" = YYYY-MM-DD)
    try:
        if desde_str:
            desde = datetime.strptime(desde_str, "%Y-%m-%d").date()
    except ValueError:
        desde = None

    try:
        if hasta_str:
            hasta = datetime.strptime(hasta_str, "%Y-%m-%d").date()
    except ValueError:
        hasta = None

    # aplica filtros
    # Si docum_fecha_emi es DateField:
    if desde:
        documentos = documentos.filter(docum_fecha_emi__gte=desde)
    if hasta:
        documentos = documentos.filter(docum_fecha_emi__lte=hasta)

    # ==============================
    # ACUMULADORES + CÁLCULOS PARA UI
    # ==============================
    total_ingresos = Decimal(0)
    total_egresos = Decimal(0)
    docs_pendientes = 0
    saldo_pendiente_ingresos = Decimal(0)
    saldo_pendiente_egresos = Decimal(0)

    for doc in documentos:
        # 1) tipo transacción (para clase visual doc-ingreso/doc-egreso)
        doc.trans_tipo = None
        for trans in doc.transacciones.all():
            if trans.tipo and trans.tipo.tipo_trans == "INGRESO":
                doc.trans_tipo = "INGRESO"
                break
            elif trans.tipo and trans.tipo.tipo_trans == "EGRESO":
                doc.trans_tipo = "EGRESO"

        # 2) total bruto y pagado (blindado si falta producto)
        doc_total = sum(
            (det.dedoc_cant or 0) * (det.producto.produ_bruto or 0)
            for det in doc.detalles.all()
            if det.producto
        )
        doc_pagado = sum(
            (det.dedoc_pagado or 0) * (det.producto.produ_bruto or 0)
            for det in doc.detalles.all()
            if det.producto
        )

        doc.total_calc = doc_total
        doc.pagado_calc = doc_pagado
        doc.pendiente_calc = max(Decimal(0), doc_total - doc_pagado)

        # (si en template usas det.total, lo seteas aquí)
        for det in doc.detalles.all():
            det.total = (det.dedoc_cant or 0) * ((det.producto.produ_bruto or 0) if det.producto else 0)

        # 3) contadores
        if doc.docum_estado in ("PENDIENTE", "MITAD"):
            docs_pendientes += 1

        # 4) acumuladores contables
        if doc.trans_tipo == "INGRESO":
            total_ingresos += doc_total
            saldo_pendiente_ingresos += doc.pendiente_calc
        elif doc.trans_tipo == "EGRESO":
            total_egresos += doc_total
            saldo_pendiente_egresos += doc.pendiente_calc

    balance = total_ingresos - total_egresos

    # IVA total gastado (asumiendo que total_egresos es BRUTO con IVA incluido)
    divisor = Decimal("1.19")
    total_gastado = total_egresos or Decimal(0)

    neto_gastado = (total_gastado / divisor) if total_gastado else Decimal(0)
    iva_total_gastado = total_gastado - neto_gastado

    # Redondeo a pesos (0 decimales) para evitar valores tipo 58.823,5294...
    neto_gastado = neto_gastado.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    iva_total_gastado = iva_total_gastado.quantize(Decimal("1"), rounding=ROUND_HALF_UP)


    return render(request, "empresa_persona/ver_persona.html", {
        "persona": persona,
        "form_p": form_p,
        "form_d": form_d,
        "proyectos": proyectos,
        "documentos": documentos,
        "desde": desde_str,   
        "hasta": hasta_str,
        "docs_pendientes": docs_pendientes,
        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
        "balance": balance,
        "saldo_pendiente_ingresos": saldo_pendiente_ingresos,
        "saldo_pendiente_egresos": saldo_pendiente_egresos,
        "proyectos_pendientes": proyectos_pendientes,
        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
        "iva_total_gastado": iva_total_gastado,
    })

#========================
#  EXPORTAR EXCEL
#========================
def export_persona_excel(request, pk):
    persona = get_object_or_404(EmpresaPersona, pk=pk)
    hoy = timezone.now().date()

    documentos = (
        Documento.objects
        .filter(empresa=persona)
        .exclude(docum_estado="ANULADO")
        .prefetch_related("detalles__producto", "transacciones__tipo")
        .order_by("docum_fecha_ven", "docum_num")
    )

    # --- helper: tipo trans ---
    def get_trans_tipo(doc):
        t = ""
        if doc.transacciones.exists() and doc.transacciones.first().tipo:
            t = (doc.transacciones.first().tipo.tipo_trans or "").upper()
        return "EGRESO" if "EGRESO" in t else "INGRESO"

    # --- preparar filas + saldos ---
    filas = []
    saldo_favor = 0   # INGRESO pendiente
    saldo_contra = 0  # EGRESO pendiente

    for doc in documentos:
        total_bruto = 0
        pagado = 0

        for det in doc.detalles.all():
            precio = int(det.producto.produ_bruto) if det.producto and det.producto.produ_bruto else 0
            total_bruto += int(det.dedoc_cant or 0) * precio
            pagado += int(det.dedoc_pagado or 0) * precio

        pendiente = max(total_bruto - pagado, 0)

        trans_tipo = get_trans_tipo(doc)
        if pendiente > 0:
            if trans_tipo == "INGRESO":
                saldo_favor += pendiente
            else:
                saldo_contra += pendiente

        neto = round(total_bruto / 1.19) if total_bruto else 0
        iva = total_bruto - neto

        dias_vencida = ""
        if doc.docum_fecha_ven and doc.docum_fecha_ven < hoy and pendiente > 0:
            dias_vencida = (hoy - doc.docum_fecha_ven).days

        filas.append({
            "doc": doc.docum_num,
            "fecha": doc.docum_fecha_emi,
            "venc": doc.docum_fecha_ven,
            "tipo": trans_tipo,
            "estado": doc.docum_estado,
            "neto": neto,
            "iva": iva,
            "bruto": total_bruto,
            "pagado": pagado,
            "pendiente": pendiente,
            "dias": dias_vencida,
        })

    # --- Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_title = PatternFill("solid", fgColor="F2F2F2")
    fill_header = PatternFill("solid", fgColor="E7EEF8")

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)

    # Título
    ws["A1"] = "Reporte deuda cliente"
    ws["A1"].font = title_font
    ws["A1"].fill = fill_title
    ws.merge_cells("A1:K1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Datos cliente
    ws["A3"] = "Cliente"
    ws["A3"].font = bold
    ws["B3"] = persona.emppe_nom or ""
    ws["D3"] = "RUT"
    ws["D3"].font = bold
    ws["E3"] = persona.emppe_rut or ""

    # Saldos
    ws["A5"] = "Saldo pendiente a favor"
    ws["A5"].font = bold
    ws["B5"] = saldo_favor

    ws["D5"] = "Saldo pendiente en contra"
    ws["D5"].font = bold
    ws["E5"] = saldo_contra

    for c in ["B5", "E5"]:
        ws[c].number_format = '#,##0'

    # Encabezados tabla (similar a tu formato)
    headers = [
        "Doc", "Fecha", "Vencimiento", "Tipo", "Estado",
        "Valor neto", "IVA", "Valor Bruto", "Pagado", "Saldo pendiente", "Días Vencida"
    ]
    start_row = 7
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = bold
        cell.fill = fill_header
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Filas
    r = start_row + 1
    for f in filas:
        ws.cell(r, 1, f["doc"]).border = border
        ws.cell(r, 2, f["fecha"].strftime("%d/%m/%Y") if f["fecha"] else "").border = border
        ws.cell(r, 3, f["venc"].strftime("%d/%m/%Y") if f["venc"] else "").border = border
        ws.cell(r, 4, f["tipo"]).border = border
        ws.cell(r, 5, f["estado"]).border = border

        for idx, key in enumerate(["neto", "iva", "bruto", "pagado", "pendiente"], start=6):
            c = ws.cell(r, idx, int(f[key] or 0))
            c.number_format = '#,##0'
            c.border = border
            c.alignment = Alignment(horizontal="right")

        ws.cell(r, 11, f["dias"]).border = border
        r += 1

    # Anchos (para que se vea “tipo reporte”)
    widths = [10, 12, 12, 10, 12, 12, 10, 14, 12, 15, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename=reporte_{persona.emppe_nom}.xlsx'
    wb.save(response)
    return response

#========================
#  EXPORTAR PDF
#========================
def export_persona_pdf(request, pk):
    persona = get_object_or_404(EmpresaPersona, pk=pk)
    hoy = timezone.now().date()

    documentos = (
        Documento.objects
        .filter(empresa=persona)
        .exclude(docum_estado="ANULADO")
        .prefetch_related("detalles__producto", "transacciones__tipo")
        .order_by("docum_fecha_ven", "docum_num")
    )

    def get_trans_tipo(doc):
        t = ""
        if doc.transacciones.exists() and doc.transacciones.first().tipo:
            t = (doc.transacciones.first().tipo.tipo_trans or "").upper()
        return "EGRESO" if "EGRESO" in t else "INGRESO"

    filas = []
    saldo_favor = 0
    saldo_contra = 0

    for doc in documentos:
        total_bruto = 0
        pagado = 0
        for det in doc.detalles.all():
            precio = int(det.producto.produ_bruto) if det.producto and det.producto.produ_bruto else 0
            total_bruto += int(det.dedoc_cant or 0) * precio
            pagado += int(det.dedoc_pagado or 0) * precio

        pendiente = max(total_bruto - pagado, 0)
        trans_tipo = get_trans_tipo(doc)

        if pendiente > 0:
            if trans_tipo == "INGRESO":
                saldo_favor += pendiente
            else:
                saldo_contra += pendiente

        dias_vencida = ""
        if doc.docum_fecha_ven and doc.docum_fecha_ven < hoy and pendiente > 0:
            dias_vencida = (hoy - doc.docum_fecha_ven).days

        filas.append((doc.docum_num, doc.docum_fecha_emi, doc.docum_fecha_ven, trans_tipo, doc.docum_estado, total_bruto, pagado, pendiente, dias_vencida))

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename=reporte_{persona.emppe_nom}.pdf'

    p = canvas.Canvas(response, pagesize=letter)
    y = 760

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Reporte deuda cliente")
    y -= 25

    p.setFont("Helvetica", 10)
    p.drawString(50, y, f"Cliente: {persona.emppe_nom or ''}   |   RUT: {persona.emppe_rut or ''}")
    y -= 16
    p.drawString(50, y, f"Saldo pendiente a favor: ${saldo_favor:,}".replace(",", "."))
    y -= 14
    p.drawString(50, y, f"Saldo pendiente en contra: ${saldo_contra:,}".replace(",", "."))
    y -= 22

    p.setFont("Helvetica-Bold", 9)
    p.drawString(50, y, "Doc")
    p.drawString(90, y, "Emisión")
    p.drawString(150, y, "Venc.")
    p.drawString(210, y, "Tipo")
    p.drawString(255, y, "Estado")
    p.drawRightString(430, y, "Bruto")
    p.drawRightString(500, y, "Pagado")
    p.drawRightString(570, y, "Pend.")
    y -= 12

    p.setFont("Helvetica", 9)
    for (num, fe, fv, tipo, estado, bruto, pagado, pend, dias) in filas:
        if y < 80:
            p.showPage()
            y = 760
            p.setFont("Helvetica", 9)

        p.drawString(50, y, str(num))
        p.drawString(90, y, fe.strftime("%d/%m/%Y") if fe else "")
        p.drawString(150, y, fv.strftime("%d/%m/%Y") if fv else "")
        p.drawString(210, y, tipo)
        p.drawString(255, y, estado)
        p.drawRightString(430, y, f"${bruto:,}".replace(",", "."))
        p.drawRightString(500, y, f"${pagado:,}".replace(",", "."))
        p.drawRightString(570, y, f"${pend:,}".replace(",", "."))
        y -= 12

    p.showPage()
    p.save()
    return response

def _clp_int(value: Decimal) -> int:
    """Redondea a CLP (entero) evitando decimales infinitos."""
    if value is None:
        return 0
    if not isinstance(value, Decimal):
        value = Decimal(str(value))
    return int(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def cc_clientes(request):
    # Base: todas las empresas/personas activas
    personas_qs = (
        EmpresaPersona.objects
        .filter(emppe_est=True)
        .order_by("emppe_nom")
    )

    # Documentos relacionados (misma carga pesada que tú ya usas en ver_persona)
    documentos_qs = (
        Documento.objects
        .select_related("tipo_doc", "proyecto", "empresa", "forma_pago", "forma_pago__tipo_pago")
        .prefetch_related("detalles__producto", "transacciones__tipo")
        .order_by("-docum_fecha_emi", "-docum_id")
    )

    # OJO: si tu related_name NO es "documentos", cambia "documentos" por "documento_set"
    personas = personas_qs.prefetch_related(
        Prefetch("documentos", queryset=documentos_qs, to_attr="docs_cc")
    )

    # Calcular métricas por persona
    for p in personas:
        docs = getattr(p, "docs_cc", [])

        p.cc_docs_count = 0
        p.cc_total_ingresado = Decimal(0)
        p.cc_saldo_pendiente_ingresos = Decimal(0)
        p.cc_total_egresado = Decimal(0)
        p.cc_saldo_pendiente_egresos = Decimal(0)

        for doc in docs:
            # si quieres excluir anulados:
            # if (doc.docum_estado or "").upper() == "ANULADO":
            #     continue

            p.cc_docs_count += 1

            bruto = sum(
                (det.dedoc_cant or 0) * (det.producto.produ_bruto or 0)
                for det in doc.detalles.all()
                if det.producto
            )

            pagado = sum(
                (det.dedoc_pagado or 0) * (det.producto.produ_bruto or 0)
                for det in doc.detalles.all()
                if det.producto
            )

            pendiente = bruto - pagado
            if pendiente < 0:
                pendiente = Decimal(0)

            # detectar ingreso/egreso por transacciones asociadas
            trans_tipo = None
            for trans in doc.transacciones.all():
                t = (trans.tipo.tipo_trans or "").upper()
                if t == "INGRESO":
                    trans_tipo = "INGRESO"
                    break
                elif t == "EGRESO":
                    trans_tipo = "EGRESO"

            if trans_tipo == "INGRESO":
                p.cc_total_ingresado += bruto
                p.cc_saldo_pendiente_ingresos += pendiente
            elif trans_tipo == "EGRESO":
                p.cc_total_egresado += bruto
                p.cc_saldo_pendiente_egresos += pendiente

    return render(request, "empresa_persona/cc_clientes.html", {
        "personas": personas
    })